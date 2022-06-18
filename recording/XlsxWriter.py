""" Functions for converting json files and objects to an Excel xlsx file """
import json
import os
from datetime import datetime

from numpy import sqrt
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.chart import Reference, Series, ScatterChart, BarChart
from openpyxl.chart.data_source import NumVal, NumDataSource, NumData
from openpyxl.chart.marker import Marker
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart.error_bar import ErrorBars

from Constants import RESULTS_DIR


def jsonFileToXlsx(jsonFileName, workbookName, sheetName, headerColor="FFAED581",
                   color1="FFF1F8E9", color2="FFDCEDC8", sep=True):
    with open(jsonFileName, 'r') as jsonFile:
        jsonData = json.load(jsonFile)

    jsonToXlsx(jsonData, workbookName, sheetName, headerColor, color1, color2, sep)


def jsonToXlsx(jsonData, workbookName, sheetName, headerColor="FFAED581",
               color1="FFF1F8E9", color2="FFDCEDC8", sep=True):
    jsonData['SIM_END_TIME'] = datetime.now()

    # Create the folder and file with the results if they do not exist, or just get them
    path = createOrGetPath(workbookName)
    workbook = createOrGetWorkbook(path)
    worksheet = createOrGetSheet(workbook, sheetName)

    row = worksheet.max_row
    if row == 1:
        headers = jsonData.keys()
        worksheet.append(["" for _ in range(len(headers) + 1)])
        worksheet.append(formatHeaderCells(worksheet, headers, headerColor))
        row += 1

    if sep:
        data = separateIntoRows(jsonData.values())
        for rowOfData in data:
            worksheet.append(formatBodyCells(worksheet, rowOfData, row, color1, color2))
            row += 1
    else:
        worksheet.append(formatBodyCells(worksheet, jsonData.values(), row, color1, color2))

    workbook.save(path)


def writeSummary(jsonData, workbookName, summarySheetName, resultsSheetName, ignore,
                 headerColor="FFAED581", color1="FFF1F8E9", color2="FFDCEDC8"):
    path = createOrGetPath(workbookName)
    workbook = createOrGetWorkbook(path)
    summarySheet = createOrGetSheet(workbook, summarySheetName)
    try:
        resultsSheet = workbook[resultsSheetName]
    except KeyError:
        workbook.close()
        return

    if summarySheet.max_row == 1:
        headers = ["Settings"] + [header for header in jsonData.keys() if header not in ignore] + \
                  [f"{header} Std Err" for header in jsonData.keys() if header not in ignore]  # Note: These values aren't used to help make the chart because openpyxl is hard to work with.
        summarySheet.append(["" for _ in range(len(headers) + 1)])
        summarySheet.append(formatHeaderCells(summarySheet, headers, headerColor))

    summaryRow = findSettingRow(summarySheet, resultsSheetName)
    color = color1 if summaryRow % 2 == 1 else color2

    lastResultRow = resultsSheet.max_row
    if lastResultRow > 2:
        # Get the columns from the results sheet that we will be referencing in the summary sheet
        columns = [i for i, header in enumerate(jsonData.keys(), 2) if header not in ignore]
        formulas = [resultsSheetName]
        stderrors = []  # Note: These values aren't used to help make the chart because openpyxl is hard to work with.
        for columnNum in range(len(columns)):
            # Convert the column from a number to a letter
            col = get_column_letter(columns[columnNum])
            # Append a formula that takes the average of the values in the next column
            formulas.append(f"=AVERAGE('{resultsSheetName}'!{col}3:{col}{lastResultRow})")
            stderrors.append(f'=IFERROR(STDEV({resultsSheetName}!{col}3:{col}{lastResultRow})/{sqrt(lastResultRow - 2)},0)')
        formulas += stderrors
        # Add each formula as a cell in the summary sheet
        for i, formula in enumerate(formulas, 2):
            summarySheet.cell(row=summaryRow, column=i).value = formula
            summarySheet.cell(row=summaryRow, column=i).fill = PatternFill(fill_type="solid", start_color=color)
            summarySheet.cell(row=summaryRow, column=i).border = Border(left=Side(style='thin'),
                                                                        right=Side(style='thin'),
                                                                        top=Side(style='thin'),
                                                                        bottom=Side(style='thin'))

    workbook.save(path)


def createOrGetPath(workbookName):
    path = f'{RESULTS_DIR}'
    path_exists = os.path.exists(path)
    if not path_exists:
        os.makedirs(path)
    if not workbookName.endswith('.xlsx'):
        workbookName += '.xlsx'
    return f'{path}/{workbookName}'


def createOrGetWorkbook(path):
    try:
        return load_workbook(path)
    except FileNotFoundError:
        return Workbook()


def createOrGetSheet(workbook, sheetName):
    try:
        return workbook[sheetName]
    except KeyError:
        return workbook.create_sheet(title=sheetName)


def formatHeaderCells(worksheet, data, headerColor):
    data = [""] + [item.title().replace('_', ' ') for item in list(data)]
    font = Font(bold=True)
    fill = PatternFill(fill_type="solid", start_color=headerColor)
    for i, c in enumerate(data, 1):
        yield formatCell(worksheet, c, i, font, fill)


def formatBodyCells(worksheet, data, row, color1, color2):
    data = [""] + list(data)
    font = Font(bold=False)
    fill = PatternFill(fill_type="solid", start_color=color1)
    if row % 2 == 1:
        fill = PatternFill(fill_type="solid", start_color=color2)
    for i, c in enumerate(data, 1):
        yield formatCell(worksheet, c, i, font, fill)


def formatCell(worksheet, c, i, font, fill):
    if i == 1:
        worksheet.column_dimensions[get_column_letter(i)].width = 4
        c = Cell(worksheet, column="A", row=1, value=c)
    else:
        if len(f"{c}") > worksheet.column_dimensions[get_column_letter(i)].width:
            worksheet.column_dimensions[get_column_letter(i)].width = len(f"{c}")
        if type(c) == list:
            c = f'{c}'
        c = Cell(worksheet, column="A", row=1, value=c)
        c.font = font
        c.fill = fill
        c.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    return c


def separateIntoRows(values):
    """ Use this when each of the values in the json file are stored as lists but should be recorded as single
    values in the xlsx file. """
    rows = []
    for value in values:
        if type(value) == list:
            if len(value) == 0:
                if len(rows) <= 0:
                    rows.append([])
                for row in rows:
                    row.append(None)
            for rowIndex, item in enumerate(value):
                if len(rows) <= rowIndex:
                    rows.append([])
                rows[rowIndex].append(item)
        else:
            if len(rows) <= 0:
                rows.append([])
            for row in rows:
                row.append(value)
    return rows


def findSettingRow(worksheet, settingName):
    for row in worksheet.iter_rows(2):
        for cell in row:
            if cell.value == settingName:
                return cell.row
    return worksheet.max_row + 1


def findSettingColumn(worksheet, settingName, columnNameRow):
    for row in worksheet.iter_rows(min_row=columnNameRow, min_col=1, max_row=columnNameRow,
                                   max_col=worksheet.max_column):
        for cell in row:
            if cell.value == settingName:
                return cell.column
    return -1


def createScatterPlot(workbookName, worksheetName, columnNameX, columnNameY, columnNameRow, chartIndex):
    path = createOrGetPath(workbookName)
    workbook = createOrGetWorkbook(path)
    worksheet = createOrGetSheet(workbook, worksheetName)

    if chartIndex > len(worksheet._charts):
        raise IndexError("The chart index must be an existing index or the next available one.\n"
                         "Current index: " + str(chartIndex) + "\n"
                         "Last existing index: " + str(len(worksheet._charts) - 1) + "\n"
                         "Next available index: " + str(len(worksheet._charts)))

    columnX = findSettingColumn(worksheet, columnNameX, columnNameRow)
    if columnX == -1:
        raise ValueError("Column name not found: " + columnNameX)
    columnY = findSettingColumn(worksheet, columnNameY, columnNameRow)
    if columnY == -1:
        raise ValueError("Column name not found: " + columnNameY)

    # The values that will make up the x and y axes.
    valuesX = Reference(worksheet, min_col=columnX, min_row=columnNameRow + 1, max_col=columnX,
                        max_row=worksheet.max_row)
    valuesY = Reference(worksheet, min_col=columnY, min_row=columnNameRow + 1, max_col=columnY,
                        max_row=worksheet.max_row)

    series = Series(valuesX, valuesY)
    series.marker = Marker('circle')
    series.graphicalProperties.line.noFill = True

    chart = ScatterChart()
    chart.series = [series]
    chart.title = f"{columnNameX} vs {columnNameY}"
    chart.x_axis.title = columnNameX
    chart.y_axis.title = columnNameY
    chart.legend = None

    if chartIndex == len(worksheet._charts):
        columnLetter = get_column_letter(worksheet.max_column + 2)
        worksheet.add_chart(chart, f"{columnLetter}{columnNameRow + chartIndex * 15}")
    else:
        columnLetter = get_column_letter(worksheet.max_column + 2)
        chart.anchor = f"{columnLetter}{columnNameRow + chartIndex * 15}"
        worksheet._charts[chartIndex] = chart

    workbook.save(path)


def createScatterPlots(workbookName, worksheetName, chartAxesList, columnNameRow, chartIndex=0):
    def setData(ch, valuesX, valuesY, i):
        series = Series(valuesY, valuesX)
        series.marker = Marker('circle')
        series.graphicalProperties.line.noFill = True
        ch.series = [series]
        ch.x_axis.scaling.min = 0

    createCharts(workbookName, worksheetName, chartAxesList, columnNameRow, ScatterChart, setData, chartIndex)


def createBarCharts(workbookName, worksheetName, chartAxesList, columnNameRow, stdErrs, chartIndex=0):
    def setData(ch, valuesX, valuesY, i):
        series = SeriesFactory(valuesY)
        series.errBars = listToErrorBars(stdErrs[i], stdErrs[i], errValType='cust')

        ch.series.append(series)
        ch.set_categories(valuesX)

    createCharts(workbookName, worksheetName, chartAxesList, columnNameRow, BarChart, setData, chartIndex)


def createCharts(workbookName, worksheetName, chartAxesList, columnNameRow, chartType, setData, chartIndex=0):
    path = createOrGetPath(workbookName)
    workbook = createOrGetWorkbook(path)
    worksheet = createOrGetSheet(workbook, worksheetName)

    if chartIndex > len(worksheet._charts):
        raise IndexError("The chart index must be an existing index or the next available one.\n"
                         "Current index: " + str(chartIndex) + "\n"
                         "Last existing index: " + str(len(worksheet._charts) - 1) + "\n"
                         "Next available index: " + str(len(worksheet._charts)))

    for i, chartAxes in enumerate(chartAxesList):
        columnX = findSettingColumn(worksheet, chartAxes[0], columnNameRow)
        if columnX == -1:
            raise ValueError(f"Column name not found: {chartAxes[0]}")
        columnY = findSettingColumn(worksheet, chartAxes[1], columnNameRow)
        if columnY == -1:
            raise ValueError(f"Column name not found: {chartAxes[1]}")

        # The values that will make up the x and y axes.
        valuesX = Reference(worksheet, min_col=columnX, min_row=columnNameRow + 1, max_col=columnX,
                            max_row=worksheet.max_row)
        valuesY = Reference(worksheet, min_col=columnY, min_row=columnNameRow + 1, max_col=columnY,
                            max_row=worksheet.max_row)

        chart = chartType()
        setData(chart, valuesX, valuesY, i)
        chart.title = f"{chartAxes[0]} vs {chartAxes[1]}"
        chart.x_axis.title = chartAxes[0]
        chart.y_axis.title = chartAxes[1]
        chart.y_axis.scaling.min = 0
        chart.legend = None

        if chartIndex == len(worksheet._charts):
            columnLetter = get_column_letter(worksheet.max_column + 2)
            worksheet.add_chart(chart, f"{columnLetter}{columnNameRow + chartIndex * 15}")
        else:
            columnLetter = get_column_letter(worksheet.max_column + 2)
            chart.anchor = f"{columnLetter}{columnNameRow + chartIndex * 15}"
            worksheet._charts[chartIndex] = chart
        chartIndex += 1

    workbook.save(path)


def listToErrorBars(plus, minus, errDir='y', errValType='stdErr'):
    """ Returns ErrorBar from lists of error values """

    # Convert to list of NumVal
    numvals_plus = [NumVal(i, None, v=x) for i, x in enumerate(plus)]
    numvals_minus = [NumVal(i, None, v=x) for i, x in enumerate(minus)]

    # Convert to NumData
    nd_plus = NumData(pt=numvals_plus)
    nd_minus = NumData(pt=numvals_minus)

    # Convert to NumDataSource
    nds_plus = NumDataSource(numLit=nd_plus)
    nds_minus = NumDataSource(numLit=nd_minus)

    return ErrorBars(errDir=errDir, errValType=errValType, plus=nds_plus, minus=nds_minus)


def getResultsStdErrs(workbookName, colNames, columnNameRow):
    path = createOrGetPath(workbookName)
    workbook = createOrGetWorkbook(path)

    resultSheetNames = [wsName for wsName in workbook.sheetnames if "Results" in wsName]
    stdErrs = [[] for _ in colNames]

    for worksheetName in resultSheetNames:
        worksheet = createOrGetSheet(workbook, worksheetName)
        for i, colName in enumerate(colNames):
            dataCol = findSettingColumn(worksheet, colName, columnNameRow)
            if dataCol == -1:
                raise ValueError(f"Column name not found: {colName}")

            numbers = [worksheet.cell(row=row, column=dataCol).internal_value
                       for row in range(columnNameRow + 1, worksheet.max_row + 1)]

            stdErrs[i].append(calcStdErr(numbers))

    workbook.close()

    return stdErrs


def calcStdErr(numbers):
    try:
        mean = sum(numbers) / len(numbers)
        variance = sum([(x - mean) ** 2 for x in numbers]) / (len(numbers) - 1)
        stdDev = sqrt(variance)
        return stdDev / sqrt(len(numbers))
    except ZeroDivisionError:
        return 0
    except TypeError:
        print(f"Type Error in calcStdErr. Probably because of numbers: {numbers}")
        return 0
